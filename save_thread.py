from work_queue import WorkQueue
from openpyxl import Workbook, load_workbook

class SaveThread:    
    TITLE_KEY = 'title'
    AUTHOR_KEY = 'author'
    ACK_KEY = 'acknowledgements'
    
    SAVE_PATH = 'thesis_contents.xlsx'

    def work(self):
        while True:
            saveData: dict = WorkQueue.saveQueue.get()              # save queue 모니터링(bloking)
            excelFile = None
            try:
                excelFile = load_workbook(SaveThread.SAVE_PATH)     # excel 존재할 경우
            except FileNotFoundError:
                excelFile = Workbook()                              # 존재하지 않을 경우 새로운 file 생성

            workSheet = excelFile.active                            # 활성화된 시트
            workSheet.title = 'thesis'                              # 시트 제목 변경

            workSheet.append([
                saveData[SaveThread.TITLE_KEY],                     # 제목
                saveData[SaveThread.AUTHOR_KEY],                    # 저자
                saveData[SaveThread.ACK_KEY],                       # 사사
            ])

            excelFile.save(SaveThread.SAVE_PATH)                    # 저장
            print('저장 완료')
